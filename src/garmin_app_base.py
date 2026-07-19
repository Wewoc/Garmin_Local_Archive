#!/usr/bin/env python3
# SPDX-License-Identifier: GPL-3.0-or-later
# Copyright (C) 2024 Wewoc (github.com/wewoc)

"""
garmin_app_base.py
Garmin Local Archive — QMainWindow Base

GarminApp — PyQt6 QMainWindow assembler. Owns shared state, log widget,
_dispatch(), and startup sequence. All panel logic lives in app/panel_*.py.

Execution-model hooks (abstract — subclass implements):
  _run(script_name, ...)  — subprocess (App) or importlib (Standalone)
  _log_bg(text)           — thread-safe log dispatch
  _is_running() -> bool   — sync running state
  _stop_collector()       — stop running sync

Worker rule (D-5): workers never touch widgets.
  All UI updates via self._dispatch(fn) or self._dispatch(lambda: ...).

v1.5.4 — PyQt6 migration
  Panels are QWidget subclasses, composed via self._panel_*.
  Mixin inheritance replaced by composition (D-1).
  pyqtSignal replaces self.after() for cross-thread dispatch (D-2).
  threading.Thread retained — QThread not introduced (D-3).
  Shared state remains on GarminApp with Owner-Matrix (D-4).
"""

import threading
from pathlib import Path

from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QHBoxLayout, QVBoxLayout,
    QLabel, QPushButton, QPlainTextEdit,
    QScrollArea, QFrame, QSizePolicy, QMessageBox,
    QTabWidget, QComboBox, QApplication,
)
from PyQt6.QtCore import Qt, QTimer, QUrl, pyqtSignal, pyqtSlot
from PyQt6.QtGui import QFont, QTextCursor
from PyQt6.QtWebEngineWidgets import QWebEngineView

import garmin_app_settings as _settings
import garmin_app_controller as _controller
import garmin_redact as _redact

from app.panel_settings   import PanelSettings
from app.panel_connection import PanelConnection
from app.panel_archive    import PanelArchive
from app.panel_timer      import PanelTimer
from app.panel_outputs    import PanelOutputs
from app.panel_home       import PanelHome

# ── Settings — re-exported for garmin_app.py / garmin_app_standalone.py ───────
SETTINGS_FILE    = _settings.SETTINGS_FILE
DEFAULT_SETTINGS = _settings.DEFAULT_SETTINGS
load_settings    = _settings.load_settings
save_password    = _settings.save_password
load_password    = _settings.load_password
delete_password  = _settings.delete_password
_open_url        = _settings._open_url

from version import APP_VERSION




class GarminApp(QMainWindow):
    """
    QMainWindow assembler for Garmin Local Archive.

    Worker rule (D-5): no panel or worker may read or write widgets
    from a background thread. All UI updates must go through _dispatch().

    State ownership (D-4):
      _ctx_running         → PanelOutputs (write) / others read
      _context_stop_event  → PanelOutputs (write) / cross-thread .set()
      _mirror_running      → PanelArchive (write) / others read
      _connection_verified → PanelConnection (write) / others read
      _timer_active        → PanelTimer (write) / others read
      _timer_stop          → PanelTimer (write) / cross-thread .set()
      _timer_generation    → PanelTimer (write) / others read
      _stopped_by_user     → PanelOutputs (write) / others read
      _last_html           → PanelOutputs (write) / others read
    """

    # ── Colors (class attributes — available to all panels via self._app.*) ────
    BG      = "#12101f"
    BG2     = "#1a1729"
    BG3     = "#231f38"
    ACCENT  = "#a259f7"
    ACCENT2 = "#6e3fcf"
    TEXT    = "#eaeaea"
    TEXT2   = "#a0a0b0"
    GREEN   = "#4ecca3"
    YELLOW  = "#f5a623"

    # ── Thread-safe dispatch signal (D-2) ──────────────────────────────────────
    # Defined at class level — emitting from any thread queues execution
    # on the Main Thread automatically (Qt queued connection).
    _dispatch_signal = pyqtSignal(object)

    def __init__(self):
        super().__init__()
        self.settings = load_settings()

        # Connect dispatch signal to slot — Qt routes cross-thread emissions
        # via queued connection automatically.
        self._dispatch_signal.connect(self._dispatch_slot)

        # ── Shared State (D-4) ─────────────────────────────────────────────────
        self._last_html           = None
        self._stopped_by_user     = False
        self._connection_verified = False
        self._timer_conn_verified = False
        self._timer_active        = False
        self._timer_stop          = threading.Event()
        self._timer_next_mode     = "repair"
        self._timer_generation    = 0
        self._ctx_running         = False
        self._context_stop_event  = threading.Event()
        self._dialog_open         = False

        # ── Window ────────────────────────────────────────────────────────────
        self.setWindowTitle("Garmin Local Archive")
        self.setMinimumSize(920, 760)
        self.resize(1100, 900)
        self.setStyleSheet(f"background: {self.BG}; color: {self.TEXT};")
        QApplication.instance().setStyleSheet(
            f"QToolTip {{ background: {self.BG3}; color: {self.TEXT}; "
            f"border: 1px solid {self.ACCENT}; padding: 4px; "
            f"font-family: 'Segoe UI'; font-size: 9pt; }}")

        # ── Build UI ──────────────────────────────────────────────────────────
        self._build_ui()

        # ── Load settings into panels ─────────────────────────────────────────
        self._panel_settings.load_settings_to_ui()
        self._panel_timer.load_timer_settings(self.settings)

        # ── Startup ───────────────────────────────────────────────────────────
        QTimer.singleShot(0,   self._check_migration)
        QTimer.singleShot(200, self._panel_archive._refresh_archive_info)
        QTimer.singleShot(500, self._startup_bg_checks)
        QTimer.singleShot(300, self._scan_dashboards)
        QTimer.singleShot(400, self._ensure_mobile_landing)

    @pyqtSlot(object)
    def _dispatch_slot(self, fn):
        """Receives callables from _dispatch_signal and executes on Main Thread."""
        fn()

    def _startup_bg_checks(self):
        """Deferred startup — ensures all panels are fully constructed."""
        threading.Thread(
            target=self._check_version, daemon=True).start()
        threading.Thread(
            target=self._panel_archive._startup_integrity_check,
            daemon=True).start()
        threading.Thread(
            target=self._panel_archive._startup_mirror_check,
            daemon=True).start()

    def _ensure_mobile_landing(self):
        """Write index.html into dashboards/ if not yet present. Main Thread."""
        try:
            s        = self._panel_settings._collect_settings()
            base_dir = s.get("base_dir", "")
            if not base_dir:
                return
            import garmin_mobile_landing as _landing
            _landing.write_index_html(base_dir)
        except Exception:
            pass

    # ── UI builder ─────────────────────────────────────────────────────────────

    def _build_ui(self):
        central = QWidget()
        central.setStyleSheet(f"background: {self.BG};")
        self.setCentralWidget(central)
        root_lay = QVBoxLayout(central)
        root_lay.setContentsMargins(0, 0, 0, 0)
        root_lay.setSpacing(0)

        # ── Header ────────────────────────────────────────────────────────────
        header = QWidget()
        header.setStyleSheet(f"background: {self.BG3};")
        header.setFixedHeight(46)
        h_lay = QHBoxLayout(header)
        h_lay.setContentsMargins(20, 0, 20, 0)
        h_lay.setSpacing(8)

        unicorn = QLabel("🦄  GARMIN LOCAL ARCHIVE")
        unicorn.setFont(QFont("Segoe UI", 13, QFont.Weight.Bold))
        unicorn.setStyleSheet(f"color: {self.TEXT}; background: transparent;")
        unicorn.setCursor(Qt.CursorShape.PointingHandCursor)
        unicorn.mousePressEvent = lambda e: self._run_extended_analysis()

        ver_lbl = QLabel(APP_VERSION)
        ver_lbl.setFont(QFont("Segoe UI", 9))
        ver_lbl.setStyleSheet(f"color: {self.TEXT2}; background: transparent;")

        tagline = QLabel("local · private · yours")
        tagline.setFont(QFont("Segoe UI", 9))
        tagline.setStyleSheet(f"color: {self.TEXT}; background: transparent;")

        link = QLabel("www.github.com/Wewoc/Garmin_Local_Archive")
        link.setFont(QFont("Segoe UI", 8))
        link.setStyleSheet(
            "color: #6ab0f5; text-decoration: underline; background: transparent;")
        link.setCursor(Qt.CursorShape.PointingHandCursor)
        link.mousePressEvent = lambda e: _open_url(
            "https://www.github.com/Wewoc/Garmin_Local_Archive")

        gpl = QLabel("GNU GPL v3")
        gpl.setFont(QFont("Segoe UI", 8))
        gpl.setStyleSheet(f"color: {self.TEXT2}; background: transparent;")

        h_lay.addWidget(unicorn)
        h_lay.addWidget(ver_lbl)
        h_lay.addWidget(tagline)
        h_lay.addStretch()
        h_lay.addWidget(link)
        h_lay.addWidget(gpl)
        root_lay.addWidget(header)

        # ── Panels ────────────────────────────────────────────────────────────
        self._panel_settings   = PanelSettings(self)
        self._panel_connection = PanelConnection(self)
        self._panel_timer      = PanelTimer(self)
        self._panel_outputs    = PanelOutputs(self)
        self._panel_archive    = PanelArchive(self)
        self._panel_home       = PanelHome(self)

        # ── Fixed top (Connection & Archive Status + Daily Actions) ───────────
        root_lay.addWidget(self._panel_home)

        # ── QTabWidget: Tab 0 Dashboard, Tab 1 Files, Tab 2 Settings ──────────
        self._right_tabs = QTabWidget()
        right_tabs = self._right_tabs
        right_tabs.setStyleSheet(
            f"QTabWidget::pane {{ border: none; background: {self.BG}; }}"
            f"QTabBar::tab {{ background: {self.BG3}; color: {self.TEXT2}; "
            f"padding: 6px 18px; border: none; font-family: 'Segoe UI'; font-size: 9pt; }}"
            f"QTabBar::tab:selected {{ background: {self.BG}; color: {self.TEXT}; "
            f"border-bottom: 2px solid {self.ACCENT}; }}"
            f"QTabBar::tab:hover {{ color: {self.TEXT}; }}")

        # ── Tab 0: Dashboard ──────────────────────────────────────────────────
        right_tabs.addTab(self._panel_home.tab_widget, "Dashboard")

        # ── Tab 1: Files ───────────────────────────────────────────────────────
        tab3_widget = QWidget()
        tab3_widget.setStyleSheet(f"background: {self.BG};")
        tab3_lay = QVBoxLayout(tab3_widget)
        tab3_lay.setContentsMargins(12, 8, 12, 8)
        tab3_lay.setSpacing(6)

        xlsx_combo_row = QHBoxLayout()
        xlsx_combo_row.setSpacing(8)
        self._xlsx_combo = QComboBox()
        self._xlsx_combo.setFont(QFont("Segoe UI", 9))
        self._xlsx_combo.setStyleSheet(
            f"QComboBox {{ background: {self.BG3}; color: {self.TEXT}; "
            f"border: none; padding: 5px 10px; }}"
            f"QComboBox::drop-down {{ border: none; }}"
            f"QComboBox QAbstractItemView {{ background: {self.BG3}; "
            f"color: {self.TEXT}; "
            f"selection-background-color: {self.ACCENT2}; }}")
        self._xlsx_combo.setSizePolicy(QSizePolicy.Policy.Expanding,
                                       QSizePolicy.Policy.Fixed)
        self._xlsx_combo.currentIndexChanged.connect(self._load_selected_xlsx)
        xlsx_combo_row.addWidget(self._xlsx_combo)
        # ▼ label for xlsx_combo — Qt6/Windows suppresses native drop-down arrow
        # when any QComboBox stylesheet is set; QLabel is reliable fallback.
        _xlsx_arrow = QLabel("▼")
        _xlsx_arrow.setFont(QFont("Segoe UI", 7))
        _xlsx_arrow.setStyleSheet(
            f"color: {self.TEXT2}; background: {self.BG3}; "
            f"padding: 0px 6px 0px 0px;")
        _xlsx_arrow.setFixedWidth(16)
        xlsx_combo_row.addWidget(_xlsx_arrow)

        self._sheet_combo = QComboBox()
        self._sheet_combo.setFont(QFont("Segoe UI", 9))
        self._sheet_combo.setStyleSheet(
            f"QComboBox {{ background: {self.BG3}; color: {self.TEXT}; "
            f"border: none; padding: 5px 10px; }}"
            f"QComboBox::drop-down {{ border: none; }}"
            f"QComboBox QAbstractItemView {{ background: {self.BG3}; "
            f"color: {self.TEXT}; "
            f"selection-background-color: {self.ACCENT2}; }}")
        self._sheet_combo.setSizePolicy(QSizePolicy.Policy.Preferred,
                                        QSizePolicy.Policy.Fixed)
        self._sheet_combo.setMinimumWidth(160)
        self._sheet_combo.setVisible(False)
        self._sheet_combo.currentIndexChanged.connect(self._on_sheet_changed)
        xlsx_combo_row.addWidget(self._sheet_combo)
        # ▼ label for sheet_combo — visibility mirrors sheet_combo (hidden for single-sheet)
        self._sheet_arrow = QLabel("▼")
        self._sheet_arrow.setFont(QFont("Segoe UI", 7))
        self._sheet_arrow.setStyleSheet(
            f"color: {self.TEXT2}; background: {self.BG3}; "
            f"padding: 0px 6px 0px 0px;")
        self._sheet_arrow.setFixedWidth(16)
        self._sheet_arrow.setVisible(False)
        xlsx_combo_row.addWidget(self._sheet_arrow)

        self._xlsx_open_btn = QPushButton("Open File")
        self._xlsx_open_btn.setFont(QFont("Segoe UI", 9))
        self._xlsx_open_btn.setStyleSheet(
            f"QPushButton {{ background: {self.BG3}; color: {self.TEXT}; "
            f"border: none; padding: 5px 14px; }}"
            f"QPushButton:hover {{ background: {self.ACCENT2}; }}")
        self._xlsx_open_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._xlsx_open_btn.clicked.connect(self._open_selected_xlsx)
        xlsx_combo_row.addWidget(self._xlsx_open_btn)
        tab3_lay.addLayout(xlsx_combo_row)

        self._xlsx_view = QWebEngineView()
        self._xlsx_view.setSizePolicy(QSizePolicy.Policy.Expanding,
                                      QSizePolicy.Policy.Expanding)
        self._xlsx_view.setStyleSheet(f"background: {self.BG};")
        tab3_lay.addWidget(self._xlsx_view)
        right_tabs.addTab(tab3_widget, "Files")

        # ── Tab 2: Settings — Zwei-Spalten-Layout ─────────────────────────────
        settings_tab = QWidget()
        settings_tab.setStyleSheet(f"background: {self.BG};")
        settings_tab_lay = QHBoxLayout(settings_tab)
        settings_tab_lay.setContentsMargins(0, 0, 0, 0)
        settings_tab_lay.setSpacing(0)

        # Links — Settings-Panel (fix 400px)
        left_scroll = QScrollArea()
        left_scroll.setWidgetResizable(True)
        left_scroll.setFixedWidth(400)
        left_scroll.setFrameShape(QFrame.Shape.NoFrame)
        left_scroll.setStyleSheet(f"background: {self.BG2};")
        left_scroll.setWidget(self._panel_settings)
        settings_tab_lay.addWidget(left_scroll)

        # Trennlinie
        divider = QFrame()
        divider.setFrameShape(QFrame.Shape.VLine)
        divider.setFixedWidth(1)
        divider.setStyleSheet(f"background: {self.BG3}; border: none;")
        settings_tab_lay.addWidget(divider)

        # Rechts — Action-Panels (flex)
        right_scroll = QScrollArea()
        right_scroll.setWidgetResizable(True)
        right_scroll.setFrameShape(QFrame.Shape.NoFrame)
        right_scroll.setStyleSheet(f"background: {self.BG};")

        right_widget = QWidget()
        right_widget.setStyleSheet(f"background: {self.BG};")
        right_lay = QVBoxLayout(right_widget)
        right_lay.setContentsMargins(0, 0, 0, 0)
        right_lay.setSpacing(0)

        right_lay.addWidget(self._panel_connection)
        right_lay.addWidget(self._panel_timer)
        right_lay.addWidget(self._panel_outputs)
        right_lay.addWidget(self._panel_archive)
        right_lay.addStretch()
        right_scroll.setWidget(right_widget)
        settings_tab_lay.addWidget(right_scroll, stretch=1)

        right_tabs.addTab(settings_tab, "Settings")

        right_tabs.currentChanged.connect(self._on_tab_changed)

        root_lay.addWidget(right_tabs, stretch=1)

        # ── Log ───────────────────────────────────────────────────────────────
        log_frame = QWidget()
        log_frame.setStyleSheet(f"background: {self.BG};")
        log_lay = QVBoxLayout(log_frame)
        log_lay.setContentsMargins(0, 0, 0, 0)
        log_lay.setSpacing(0)

        log_bar = QWidget()
        log_bar.setStyleSheet(f"background: {self.BG3};")
        log_bar.setFixedHeight(28)
        log_bar_lay = QHBoxLayout(log_bar)
        log_bar_lay.setContentsMargins(12, 0, 12, 0)
        log_hdr = QLabel("LOG")
        log_hdr.setFont(QFont("Segoe UI", 7, QFont.Weight.Bold))
        log_hdr.setStyleSheet(f"color: {self.ACCENT}; background: transparent;")
        clear_btn = QPushButton("Clear")
        clear_btn.setFont(QFont("Segoe UI", 7))
        clear_btn.setStyleSheet(
            f"QPushButton {{ background: transparent; color: {self.TEXT2}; "
            f"border: none; padding: 2px 6px; }}"
            f"QPushButton:hover {{ color: {self.TEXT}; }}")
        clear_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        clear_btn.clicked.connect(self._clear_log)
        log_bar_lay.addWidget(log_hdr)
        log_bar_lay.addStretch()
        log_bar_lay.addWidget(clear_btn)
        log_lay.addWidget(log_bar)

        self.log = QPlainTextEdit()
        self.log.setReadOnly(True)
        self.log.setFont(QFont("Consolas", 8))
        self.log.setStyleSheet(
            f"QPlainTextEdit {{ background: #0a0a1a; color: {self.GREEN}; "
            f"border: none; padding: 4px; }}")
        self.log.setFixedHeight(90)
        log_lay.addWidget(self.log)
        root_lay.addWidget(log_frame)

    # ── Dispatch — thread-safe UI update via pyqtSignal ───────────────────────

    def _dispatch(self, fn, *args):
        """Schedule fn(*args) on the Main Thread. Safe to call from any thread.
        Qt routes the signal emission via queued connection automatically."""
        if args:
            self._dispatch_signal.emit(lambda: fn(*args))
        else:
            self._dispatch_signal.emit(fn)

    # ── Log helpers ────────────────────────────────────────────────────────────

    def _log(self, text: str):
        """Write to log widget. Main Thread only."""
        if not hasattr(self, "log"):
            return
        self.log.appendPlainText(_redact.redact(text))
        self.log.moveCursor(QTextCursor.MoveOperation.End)

    def _log_bg(self, text: str):
        """Thread-safe log write — dispatches to Main Thread."""
        self._dispatch(self._log, text)

    def _clear_log(self):
        self.log.clear()

    # ── Settings passthrough ───────────────────────────────────────────────────

    def _collect_settings(self) -> dict:
        """Central settings reader — delegates to PanelSettings.
        Timer fields are merged from PanelTimer. Called by Controller and panels."""
        s = self._panel_settings._collect_settings()
        s.update(self._panel_timer.get_timer_settings())
        return s

    def _safe_save(self, s: dict = None):
        self._panel_settings._safe_save(s)

    # ── Execution-model hooks (abstract) ───────────────────────────────────────

    def _run(self, script_name: str, enable_stop: bool = False,
             on_success=None, refresh_failed: bool = False,
             on_done=None, log_prefix: str = "garmin",
             env_overrides: dict = None, stop_event: threading.Event = None,
             days_left: int = None):
        """Subclass implements: subprocess (App) or importlib (Standalone)."""
        raise NotImplementedError

    def _is_running(self) -> bool:
        """Returns True if a sync is currently running. Subclass implements."""
        raise NotImplementedError

    def _stop_collector(self):
        """Stop running sync. Subclass implements."""
        raise NotImplementedError

    # ── ENV builder ────────────────────────────────────────────────────────────

    def _build_env_dict(self, s: dict, refresh_failed: bool = False) -> dict:
        return _controller.build_env_dict(s, refresh_failed)

    # ── Migration ──────────────────────────────────────────────────────────────

    def _check_migration(self):
        s        = self._collect_settings()
        base_dir = Path(s.get("base_dir", "")).expanduser()
        if not _controller.check_migration_needed(base_dir):
            return

        answer = QMessageBox.question(
            self, "Structure migration required",
            f"Old folder structure detected in:\n{base_dir}\n\n"
            f"The folders raw/, summary/ and log/ will be moved to\n"
            f"garmin_data/.\n\n"
            f"⚠ Recommendation: Create a manual backup of:\n"
            f"{base_dir}\n\n"
            f"Migrate now?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if answer != QMessageBox.StandardButton.Yes:
            QMessageBox.information(
                self, "App blocked",
                "The app cannot be used without migration.\n"
                "Please restart the app after creating a manual backup.",
            )
            self.close()
            return

        result = _controller.run_migration(base_dir)
        if result == "ok":
            QMessageBox.information(
                self, "Migration complete",
                f"Folders successfully moved to:\n{base_dir / 'garmin_data'}",
            )
        else:
            QMessageBox.critical(
                self, "Migration error",
                "Error moving folders.\n\n"
                "Please migrate manually and restart the app.",
            )
            self.close()

    # ── Version check ──────────────────────────────────────────────────────────

    def _check_version(self):
        import urllib.request
        import json as _json
        url = ("https://api.github.com/repos/Wewoc/"
               "Garmin_Local_Archive/releases/latest")
        try:
            req = urllib.request.Request(
                url, headers={"User-Agent": "GarminLocalArchive"})
            with urllib.request.urlopen(req, timeout=5) as resp:
                data = _json.loads(resp.read().decode())
            latest = data.get("tag_name", "").strip()
            title  = data.get("name", "").strip() or latest
            if latest and latest.lstrip("vV") != APP_VERSION.lstrip("vV"):
                self._dispatch(self._show_update_popup, latest, title)
        except Exception:
            pass

    def _show_update_popup(self, latest: str, title: str):
        import webbrowser
        dlg = QMessageBox(self)
        dlg.setWindowTitle("Update Available")
        dlg.setText(
            f"A new version is available: {title}\n"
            f"You are running: {APP_VERSION}"
        )
        dlg.setStyleSheet(f"background: {self.BG}; color: {self.TEXT};")
        open_btn  = dlg.addButton("Open GitHub",
                                   QMessageBox.ButtonRole.AcceptRole)
        dlg.addButton("Dismiss", QMessageBox.ButtonRole.RejectRole)
        dlg.exec()
        try:
            if dlg.clickedButton() == open_btn:
                webbrowser.open(
                    "https://github.com/Wewoc/Garmin_Local_Archive/releases/latest")
        except RuntimeError:
            pass

    # ── Extended Analysis (Easter Egg) ─────────────────────────────────────────

    def _run_extended_analysis(self):
        """Subclass (garmin_app.py) overrides with _find_python() access."""
        pass

    # ── Tab switch ────────────────────────────────────────────────────────────

    def _on_tab_changed(self, index: int):
        """Refresh Files tab on switch — catches XLSX files built since startup."""
        if index == 1:
            self._scan_xlsx_files()

    # ── Dashboard tab helpers ──────────────────────────────────────────────────

    def _scan_dashboards(self, auto_load: str = None):
        """Scan dashboards/ for HTML files and populate Dashboard combo.
        Called on startup and after every dashboard build (on_done via _dispatch)."""
        s        = self._panel_settings._collect_settings()
        dash_dir = Path(s.get("base_dir", "")) / "dashboards"
        html_files = sorted(
            dash_dir.glob("*.html"),
            key=lambda f: f.stat().st_mtime,
            reverse=True,
        ) if dash_dir.exists() else []

        combo = self._panel_home._dash_combo
        combo.blockSignals(True)
        combo.clear()
        if not html_files:
            combo.addItem("— no dashboards found —")
            combo.setEnabled(False)
        else:
            combo.setEnabled(True)
            for f in html_files:
                combo.addItem(f.name, userData=str(f))
        combo.blockSignals(False)

        if auto_load and html_files:
            paths = [str(f) for f in html_files]
            idx   = paths.index(auto_load) if auto_load in paths else 0
            combo.setCurrentIndex(idx)
            self._load_selected_dashboard()
        elif html_files:
            # Prefer health_garmin dashboard as default; fall back to index 0
            # Exclude mobile variant (health_garmin_mobile.html)
            names = [f.name for f in html_files]
            default_idx = next(
                (i for i, n in enumerate(names)
                 if "health_garmin" in n and "mobile" not in n), 0)
            combo.setCurrentIndex(default_idx)
            self._load_selected_dashboard()

    def _load_selected_dashboard(self):
        """Load the currently selected HTML file into the Dashboard QWebEngineView."""
        path = self._panel_home._dash_combo.currentData()
        if path and Path(path).exists():
            self._panel_home._dash_view.setUrl(QUrl.fromLocalFile(path))

    # ── Files tab helpers ──────────────────────────────────────────────────────

    def _scan_xlsx_files(self):
        """Scan dashboards/ for XLSX files and populate Tab 3 combo.
        Called on tab switch and after every dashboard build."""
        s         = self._panel_settings._collect_settings()
        dash_dir  = Path(s.get("base_dir", "")) / "dashboards"
        xlsx_files = sorted(
            dash_dir.glob("*.xlsx"),
            key=lambda f: f.stat().st_mtime,
            reverse=True,
        ) if dash_dir.exists() else []

        self._xlsx_combo.blockSignals(True)
        self._xlsx_combo.clear()
        if not xlsx_files:
            self._xlsx_combo.addItem("— no Excel files found —")
            self._xlsx_combo.setEnabled(False)
            self._xlsx_open_btn.setEnabled(False)
        else:
            self._xlsx_combo.setEnabled(True)
            self._xlsx_open_btn.setEnabled(True)
            for f in xlsx_files:
                self._xlsx_combo.addItem(f.name, userData=str(f))
        self._xlsx_combo.blockSignals(False)

        if xlsx_files:
            # Prefer overview_garmin as default; fall back to index 0
            names = [f.name for f in xlsx_files]
            default_idx = next(
                (i for i, n in enumerate(names) if "overview_garmin" in n), 0)
            self._xlsx_combo.setCurrentIndex(default_idx)
            self._load_selected_xlsx()

    def _load_selected_xlsx(self):
        """Read selected XLSX, populate sheet combo, render active sheet."""
        path = self._xlsx_combo.currentData()
        if not path or not Path(path).exists():
            self._sheet_combo.setVisible(False)
            return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            wb.close()
        except Exception as exc:
            self._sheet_combo.setVisible(False)
            self._xlsx_view.setHtml(
                f"<body style='background:{self.BG};color:{self.TEXT};"
                f"font-family:Segoe UI;padding:20px;'>"
                f"<b>Could not read file:</b> {exc}</body>",
                QUrl("about:blank"),
            )
            return

        # Filter chart sheets — openpyxl reads them as empty, not useful in viewer
        data_sheets = [n for n in sheet_names if not n.endswith(" - Chart")]
        if not data_sheets:
            data_sheets = sheet_names  # fallback: show all if none pass filter

        # Populate sheet combo — block signals to avoid double render
        self._sheet_combo.blockSignals(True)
        self._sheet_combo.clear()
        for name in data_sheets:
            self._sheet_combo.addItem(name)
        self._sheet_combo.setCurrentIndex(0)
        self._sheet_combo.setVisible(len(data_sheets) > 1)
        self._sheet_arrow.setVisible(len(data_sheets) > 1)
        self._sheet_combo.blockSignals(False)

        self._render_sheet(path, data_sheets[0])

    def _on_sheet_changed(self, index: int):
        """Render the newly selected sheet."""
        path = self._xlsx_combo.currentData()
        if not path or not Path(path).exists():
            return
        sheet_name = self._sheet_combo.currentText()
        if sheet_name:
            self._render_sheet(path, sheet_name)

    def _render_sheet(self, path: str, sheet_name: str):
        """Read one sheet from XLSX, write to temp file, load via setUrl().
        setUrl() replaces content atomically — avoids setHtml() stale-render
        on large sheets."""
        import tempfile
        try:
            import openpyxl
            wb   = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws   = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception as exc:
            self._xlsx_view.setHtml(
                f"<body style='background:{self.BG};color:{self.TEXT};"
                f"font-family:Segoe UI;padding:20px;'>"
                f"<b>Could not read sheet '{sheet_name}':</b> {exc}</body>",
                QUrl("about:blank"),
            )
            return

        if not rows:
            self._xlsx_view.setHtml(
                f"<body style='background:{self.BG};color:{self.TEXT};"
                f"font-family:Segoe UI;padding:20px;'>Sheet is empty.</body>",
                QUrl("about:blank"),
            )
            return

        th_style = (f"background:{self.BG3};color:{self.ACCENT};"
                    f"padding:6px 12px;text-align:left;font-size:11px;"
                    f"font-weight:600;border-bottom:1px solid {self.ACCENT};")
        td_style = (f"padding:5px 12px;font-size:11px;color:{self.TEXT};"
                    f"border-bottom:1px solid {self.BG3};")
        td_alt   = (f"padding:5px 12px;font-size:11px;color:{self.TEXT};"
                    f"border-bottom:1px solid {self.BG3};"
                    f"background:{self.BG2};")

        header = rows[0]
        body   = rows[1:]
        th_cells = "".join(
            f"<th style='{th_style}'>{(v if v is not None else '')}</th>"
            for v in header
        )
        # Detect single-char columns (e.g. sleep phase bar: D/L/R/A)
        # by scanning all body rows — gives them a narrow fixed width in HTML.
        num_cols = max((len(r) for r in body), default=0)
        _SINGLE_CHAR = set()
        for col_idx in range(num_cols):
            vals = [
                str(r[col_idx]) for r in body
                if col_idx < len(r) and r[col_idx] is not None
            ]
            if vals and all(len(v) == 1 for v in vals):
                _SINGLE_CHAR.add(col_idx)

        td_narrow = (f"padding:2px 1px;font-size:9px;font-weight:bold;"
                     f"text-align:center;width:10px;max-width:10px;"
                     f"overflow:hidden;color:{self.TEXT};"
                     f"border-bottom:1px solid {self.BG3};")
        td_narrow_alt = (f"padding:2px 1px;font-size:9px;font-weight:bold;"
                         f"text-align:center;width:10px;max-width:10px;"
                         f"overflow:hidden;color:{self.TEXT};"
                         f"border-bottom:1px solid {self.BG3};"
                         f"background:{self.BG2};")

        tr_rows = ""
        for i, row in enumerate(body):
            is_alt = i % 2
            cells = ""
            for col_idx, v in enumerate(row):
                val = v if v is not None else ""
                if col_idx in _SINGLE_CHAR:
                    style = td_narrow_alt if is_alt else td_narrow
                else:
                    style = td_alt if is_alt else td_style
                cells += f"<td style='{style}'>{val}</td>"
            tr_rows += f"<tr>{cells}</tr>"

        html = (
            f"<!DOCTYPE html><html><head><meta charset='UTF-8'>"
            f"<style>body{{background:{self.BG};margin:0;padding:12px;"
            f"font-family:'Segoe UI',sans-serif;}}"
            f"table{{border-collapse:collapse;width:100%;}}"
            f"</style></head><body>"
            f"<table><thead><tr>{th_cells}</tr></thead>"
            f"<tbody>{tr_rows}</tbody></table></body></html>"
        )

        tmp = Path(tempfile.gettempdir()) / "gla_xlsx_view.html"
        tmp.write_text(html, encoding="utf-8")
        self._xlsx_view.setUrl(QUrl.fromLocalFile(str(tmp)))

    def _open_selected_xlsx(self):
        """Open the selected XLSX in the system default application."""
        import os
        path = self._xlsx_combo.currentData()
        if path and Path(path).exists():
            os.startfile(path)

    # ── Close ──────────────────────────────────────────────────────────────────

    def closeEvent(self, event):
        """D-9: set all stop events, join threads with timeout, save settings."""
        self._timer_generation += 1
        self._timer_stop.set()
        self._context_stop_event.set()

        for t in threading.enumerate():
            if t.daemon and t != threading.main_thread():
                t.join(timeout=2.0)

        try:
            # Guard: widgets may already be deleted during pytest-qt teardown.
            # _collect_settings() reads from QLineEdit widgets — raises RuntimeError
            # if they have been deleted. Skip save in that case to avoid overwriting
            # the real settings file with empty values.
            self.settings = self._collect_settings()
            save_password(self.settings.get("password", ""))
            self._safe_save(self.settings)
        except RuntimeError:
            pass
        super().closeEvent(event)
