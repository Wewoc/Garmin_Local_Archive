#!/usr/bin/env python3
"""
dash_plotter_excel.py

Excel plotter — renders a specialist data dict to an .xlsx file.
One data sheet + one chart sheet per field.

Rules:
- No knowledge of Garmin internals, field names, or data sources.
- Fetches all design assets from dash_layout.
- Receives neutral dict from dash_runner, writes output file.

Interface:
    render(data: dict, output_path: Path, settings: dict) -> None
"""

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
import dash_layout as layout

# ══════════════════════════════════════════════════════════════════════════════
#  Shared styles — built once from dash_layout tokens
# ══════════════════════════════════════════════════════════════════════════════

_HEADER_FILL = PatternFill("solid", start_color=layout.EXCEL_HEADER_COLOR)
_HEADER_FONT = Font(name="Arial", bold=True,  color=layout.EXCEL_HEADER_FONT, size=10)
_DATA_FONT   = Font(name="Arial", bold=False, color="000000",                 size=10)
_DATE_FONT   = Font(name="Arial", bold=True,  color="000000",                 size=10)
_BORDER_SIDE = Side(style="thin", color=layout.EXCEL_BORDER_COLOR)
_CELL_BORDER = Border(
    left=_BORDER_SIDE, right=_BORDER_SIDE,
    top=_BORDER_SIDE,  bottom=_BORDER_SIDE,
)


# ══════════════════════════════════════════════════════════════════════════════
#  Internal builders
# ══════════════════════════════════════════════════════════════════════════════

def _write_field_sheets(wb: Workbook, entry: dict, date_from: str, date_to: str) -> None:
    """Write one data sheet + one chart sheet for a single field."""
    field  = entry["field"]
    series = entry.get("series") or []
    meta   = layout.get_metric_meta(field)
    label  = meta.get("label", field)
    unit   = meta.get("unit", "")
    color  = meta.get("color", "#888888").lstrip("#")
    fill   = layout.get_excel_row_color(field)

    data_title  = f"{label} - Data"
    chart_title = f"{label} - Chart"

    # ── Data sheet ────────────────────────────────────────────────────────────
    ws = wb.create_sheet(data_title)
    ws.freeze_panes = "A2"

    headers = ["Date", "Time", unit or "Value"]
    widths  = [12, 8, 10]

    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        cell            = ws.cell(1, col, h)
        cell.font       = _HEADER_FONT
        cell.fill       = _HEADER_FILL
        cell.alignment  = Alignment(horizontal="center")
        cell.border     = _CELL_BORDER
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 24
    row_fill = PatternFill("solid", start_color=fill)

    for row_idx, point in enumerate(series, start=2):
        ts  = point.get("ts", "")
        val = point.get("value")

        # Split ISO timestamp into date + time parts
        if "T" in ts:
            day, time = ts[:10], ts[11:16]
        else:
            day, time = ts, ""

        ws.cell(row_idx, 1, day).font  = _DATE_FONT
        ws.cell(row_idx, 1).alignment  = Alignment(horizontal="center")
        ws.cell(row_idx, 1).border     = _CELL_BORDER
        ws.cell(row_idx, 2, time).font = _DATA_FONT
        ws.cell(row_idx, 2).alignment  = Alignment(horizontal="center")
        ws.cell(row_idx, 2).border     = _CELL_BORDER
        ws.cell(row_idx, 3, val).font  = _DATA_FONT
        ws.cell(row_idx, 3).fill       = row_fill
        ws.cell(row_idx, 3).alignment  = Alignment(horizontal="center")
        ws.cell(row_idx, 3).border     = _CELL_BORDER
        if isinstance(val, float):
            ws.cell(row_idx, 3).number_format = "0.0"

    # ── Chart sheet ───────────────────────────────────────────────────────────
    wc    = wb.create_sheet(chart_title)
    chart = LineChart()
    chart.title        = f"{label} ({date_from} – {date_to})"
    chart.style        = 10
    chart.y_axis.title = unit or "Value"
    chart.x_axis.title = "Measurement"
    chart.width        = 30
    chart.height       = 18

    n = len(series)
    if n > 0:
        data_ref = Reference(ws, min_col=3, min_row=1, max_row=n + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.series[0].graphicalProperties.line.solidFill = color
        chart.series[0].graphicalProperties.line.width     = 15000  # 1.5pt EMU

    wc.add_chart(chart, "A1")


# ══════════════════════════════════════════════════════════════════════════════
#  Public interface
# ══════════════════════════════════════════════════════════════════════════════

# ── Group colors for Overview mode ────────────────────────────────────────────
_GROUP_COLORS = {
    "sleep":     "DDEEFF",
    "heartrate": "FFE8CC",
    "stress":    "E8FFE8",
    "day":       "FFF8CC",
    "training":  "F0E8FF",
}


# ── Phase bar color constants ─────────────────────────────────────────────────
_PHASE_BAR_COLORS = {
    "deep":  "2d6a9f",
    "light": "7eb8d4",
    "rem":   "9b7fc7",
    "awake": "d4c5a9",
}

# ── Qualifier badge colors (font, fill) ───────────────────────────────────────
_QUALIFIER_XL = {
    "EXCELLENT": ("1a7a4a", "d4f5e5"),
    "GOOD":      ("1a6a6a", "d4f0f0"),
    "FAIR":      ("8a6a00", "fff3cc"),
    "POOR":      ("8a2000", "ffe0d0"),
}

_PHASE_COUNT = 20  # number of narrow cells representing the phase bar


def _hsl_to_hex(norm: float, higher_better: bool = True) -> str:
    """
    Map normalized 0–1 value to a hex color string via HSL interpolation.
    0=red (#cc3300), 0.5=amber (#cc8800), 1=green (#2d8c00).
    """
    if higher_better:
        hue = norm * 120
    else:
        hue = (1.0 - norm) * 120
    # Approximate HSL(hue, 65%, 35%) → hex without external libs
    # Using precomputed anchor points and linear interpolation
    anchors = [
        (0,   "cc3300"),
        (30,  "cc6600"),
        (60,  "cc9900"),
        (90,  "66aa00"),
        (120, "2d8c00"),
    ]
    hue = max(0.0, min(120.0, hue))
    for i in range(len(anchors) - 1):
        h0, c0 = anchors[i]
        h1, c1 = anchors[i + 1]
        if h0 <= hue <= h1:
            t  = (hue - h0) / (h1 - h0)
            r0, g0, b0 = int(c0[0:2], 16), int(c0[2:4], 16), int(c0[4:6], 16)
            r1, g1, b1 = int(c1[0:2], 16), int(c1[2:4], 16), int(c1[4:6], 16)
            r  = int(r0 + (r1 - r0) * t)
            g  = int(g0 + (g1 - g0) * t)
            b  = int(b0 + (b1 - b0) * t)
            return f"{r:02x}{g:02x}{b:02x}"
    return "888888"


def _value_color_hex(value, low, high, higher_better=True) -> str:
    """Return hex color for a numeric value within [low, high]."""
    if value is None or high <= low:
        return "888888"
    norm = max(0.0, min(1.0, (value - low) / (high - low)))
    return _hsl_to_hex(norm, higher_better)


def _write_sleep_sheet(wb: Workbook, data: dict) -> None:
    """
    Write Sleep Dashboard as one sheet — one row per night.
    Phase bar: 20 narrow colored cells per row.
    Colored numbers: font color from HSL interpolation.
    """
    rows  = data.get("rows", [])
    refs  = data.get("refs", {})
    title = data.get("title", "Sleep Dashboard")

    ref_hrv_low,  ref_hrv_high  = refs.get("hrv_last_night",   (30, 80))
    ref_slp_low,  ref_slp_high  = refs.get("sleep_duration",   (7.0, 9.0))
    ref_bb_low,   ref_bb_high   = refs.get("body_battery_max", (50, 100))

    ws = wb.create_sheet(title[:31])
    ws.freeze_panes = "A2"

    # ── Column layout ─────────────────────────────────────────────────────────
    # Col 1:       Date
    # Col 2–21:    Phase bar (20 narrow cells)
    # Col 22:      Duration
    # Col 23:      Score
    # Col 24:      Quality
    # Col 25:      Feedback
    # Col 26:      HRV       (separator: left border)
    # Col 27:      Body Battery

    BAR_START = 2
    BAR_END   = BAR_START + _PHASE_COUNT - 1   # = 21
    COL_DUR   = BAR_END + 1                    # = 22
    COL_SCORE = COL_DUR  + 1                   # = 23
    COL_QUAL  = COL_SCORE + 1                  # = 24
    COL_FB    = COL_QUAL  + 1                  # = 25
    COL_HRV   = COL_FB    + 1                  # = 26
    COL_BB    = COL_HRV   + 1                  # = 27

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 12
    for col in range(BAR_START, BAR_END + 1):
        ws.column_dimensions[get_column_letter(col)].width = 1.5
    ws.column_dimensions[get_column_letter(COL_DUR)].width   = 9
    ws.column_dimensions[get_column_letter(COL_SCORE)].width = 7
    ws.column_dimensions[get_column_letter(COL_QUAL)].width  = 12
    ws.column_dimensions[get_column_letter(COL_FB)].width    = 32
    ws.column_dimensions[get_column_letter(COL_HRV)].width   = 8
    ws.column_dimensions[get_column_letter(COL_BB)].width    = 12

    # ── Header row ────────────────────────────────────────────────────────────
    headers = {
        1:        "Date",
        BAR_START:"Sleep Phases",
        COL_DUR:  "Duration",
        COL_SCORE:"Score",
        COL_QUAL: "Quality",
        COL_FB:   "Feedback",
        COL_HRV:  "HRV",
        COL_BB:   "Body Battery",
    }
    for col, label in headers.items():
        cell           = ws.cell(1, col, label)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border    = _CELL_BORDER
    # Merge phase bar header across 20 cells
    ws.merge_cells(
        start_row=1, start_column=BAR_START,
        end_row=1,   end_column=BAR_END
    )
    ws.row_dimensions[1].height = 24

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, row in enumerate(rows, start=2):
        date     = row.get("date", "")
        duration = row.get("duration_h")
        score    = row.get("score")
        qualifier= row.get("qualifier") or ""
        feedback = row.get("feedback") or ""
        hrv      = row.get("hrv")
        bb       = row.get("body_battery")

        ws.row_dimensions[row_idx].height = 16

        # Date
        c = ws.cell(row_idx, 1, date)
        c.font      = _DATE_FONT
        c.alignment = Alignment(horizontal="center")
        c.border    = _CELL_BORDER

        # Phase bar — 20 narrow cells, colored by phase
        phases = [
            ("deep",  row.get("deep")),
            ("light", row.get("light")),
            ("rem",   row.get("rem")),
            ("awake", row.get("awake")),
        ]
        total = sum(v for _, v in phases if v is not None)
        col   = BAR_START
        if total > 0:
            for phase_key, val in phases:
                if val is None or val <= 0:
                    continue
                cells_for_phase = max(1, round(val / total * _PHASE_COUNT))
                color = _PHASE_BAR_COLORS[phase_key]
                for _ in range(cells_for_phase):
                    if col > BAR_END:
                        break
                    c = ws.cell(row_idx, col)
                    c.fill = PatternFill("solid", start_color=color)
                    col += 1
        # Fill remaining bar cells (rounding gap) with last phase color or grey
        while col <= BAR_END:
            c = ws.cell(row_idx, col)
            c.fill = PatternFill("solid", start_color="444444")
            col += 1

        # Duration
        dur_hex = _value_color_hex(duration, ref_slp_low, ref_slp_high)
        c = ws.cell(row_idx, COL_DUR, duration)
        c.font           = Font(name="Arial", bold=True, color=dur_hex, size=10)
        c.alignment      = Alignment(horizontal="center")
        c.border         = _CELL_BORDER
        c.number_format  = "0.0"

        # Score
        score_hex = _value_color_hex(score, 30, 100)
        c = ws.cell(row_idx, COL_SCORE, score)
        c.font      = Font(name="Arial", bold=True, color=score_hex, size=10)
        c.alignment = Alignment(horizontal="center")
        c.border    = _CELL_BORDER

        # Quality badge (text + background color)
        fg_hex, bg_hex = _QUALIFIER_XL.get(qualifier, ("555555", "eeeeee"))
        c = ws.cell(row_idx, COL_QUAL, qualifier or "—")
        c.font      = Font(name="Arial", bold=True, color=fg_hex, size=10)
        c.fill      = PatternFill("solid", start_color=bg_hex)
        c.alignment = Alignment(horizontal="center")
        c.border    = _CELL_BORDER

        # Feedback (plain text, cleaned)
        cleaned_fb = feedback.replace("NEGATIVE_", "").replace("POSITIVE_", "")
        parts      = [p.capitalize().replace("_", " ") for p in cleaned_fb.split("_AND_")]
        fb_text    = " · ".join(parts) if parts and parts[0] else "—"
        c = ws.cell(row_idx, COL_FB, fb_text)
        c.font      = Font(name="Arial", bold=False, color="888888", size=10)
        c.alignment = Alignment(horizontal="left")
        c.border    = _CELL_BORDER

        # HRV (left separator border)
        hrv_hex = _value_color_hex(hrv, ref_hrv_low, ref_hrv_high)
        sep_border = Border(
            left  = Side(style="medium", color="2d6a9f"),
            right = _BORDER_SIDE, top = _BORDER_SIDE, bottom = _BORDER_SIDE,
        )
        c = ws.cell(row_idx, COL_HRV, hrv)
        c.font      = Font(name="Arial", bold=True, color=hrv_hex, size=10)
        c.alignment = Alignment(horizontal="center")
        c.border    = sep_border

        # Body Battery
        bb_hex = _value_color_hex(bb, ref_bb_low, ref_bb_high)
        c = ws.cell(row_idx, COL_BB, bb)
        c.font      = Font(name="Arial", bold=True, color=bb_hex, size=10)
        c.alignment = Alignment(horizontal="center")
        c.border    = _CELL_BORDER


def _write_overview_sheet(wb: Workbook, data: dict) -> None:
    """Write broad daily overview table — one row per day, one col per field."""
    columns   = data.get("columns", [])
    rows      = data.get("rows", [])
    title     = data.get("title", "Overview")

    ws = wb.create_sheet(title[:31])  # sheet name max 31 chars
    ws.freeze_panes = "B2"

    # Header — Date + one col per field
    ws.cell(1, 1, "Date").font      = _HEADER_FONT
    ws.cell(1, 1).fill              = _HEADER_FILL
    ws.cell(1, 1).alignment         = Alignment(horizontal="center")
    ws.cell(1, 1).border            = _CELL_BORDER
    ws.column_dimensions["A"].width = 12

    for col_idx, col in enumerate(columns, start=2):
        cell = ws.cell(1, col_idx, col["label"])
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border    = _CELL_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(len(col["label"]) + 2, 20))

    ws.row_dimensions[1].height = 36

    for row_idx, row in enumerate(rows, start=2):
        date_cell = ws.cell(row_idx, 1, row["date"])
        date_cell.font      = _DATE_FONT
        date_cell.alignment = Alignment(horizontal="center")
        date_cell.border    = _CELL_BORDER

        for col_idx, col in enumerate(columns, start=2):
            val   = row["values"].get(col["field"])
            color = _GROUP_COLORS.get(col.get("group", ""), "FFFFFF")
            cell  = ws.cell(row_idx, col_idx, val)
            cell.font      = _DATA_FONT
            cell.fill      = PatternFill("solid", start_color=color)
            cell.alignment = Alignment(horizontal="center")
            cell.border    = _CELL_BORDER
            if isinstance(val, float):
                cell.number_format = "0.00"
            elif isinstance(val, int):
                cell.number_format = "#,##0"


def render(data: dict, output_path: Path, settings: dict) -> None:
    """
    Render specialist data dict to Excel file.

    Detects output mode from dict structure:
    - "rows" present          → Overview: broad flat table
    - "fields" with "series"  → Timeseries: per-field data + chart sheets
    - "fields" with "days"    → Analysis: per-field data + chart sheets

    Args:
        data:        Dict from specialist.build().
        output_path: Full path for the output .xlsx file.
        settings:    Settings dict from GUI (reserved for future use).

    Raises:
        ValueError: if no usable data is present.
        OSError:    if output file cannot be written.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    # ── Sleep Dashboard mode ──────────────────────────────────────────────────
    if data.get("layout") == "sleep":
        if not data.get("rows"):
            raise ValueError("render: no rows in data — nothing to render")
        _write_sleep_sheet(wb, data)
        wb.save(output_path)
        return

    # ── Overview mode ─────────────────────────────────────────────────────────
    if "rows" in data:
        if not data.get("rows"):
            raise ValueError("render: no rows in data — nothing to render")
        _write_overview_sheet(wb, data)
        wb.save(output_path)
        return

    # ── Field mode (Timeseries or Analysis) ───────────────────────────────────
    fields = [f for f in data.get("fields", []) if f.get("series") or f.get("days")]
    if not fields:
        raise ValueError("render: no fields with data — nothing to render")

    date_from = data.get("date_from", "")
    date_to   = data.get("date_to", "")

    for entry in fields:
        # Normalise Analysis days → series-like list for _write_field_sheets
        if "days" in entry and "series" not in entry:
            entry = {
                **entry,
                "series": [
                    {"ts": d["date"], "value": d["value"]}
                    for d in entry["days"]
                    if d["value"] is not None
                ],
            }
        _write_field_sheets(wb, entry, date_from, date_to)

    wb.save(output_path)