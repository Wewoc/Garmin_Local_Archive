#!/usr/bin/env python3
"""
garmin_timeseries_excel.py

Reads raw JSON files and exports full intraday timeseries data to Excel.
Per metric: one sheet with raw data table, one sheet with a line chart.

Metrics covered: Heart Rate, Stress, SpO2, Body Battery, Respiration

Configuration via environment variables (all optional — hardcoded fallbacks below):
  GARMIN_OUTPUT_DIR         Root data folder (raw/ lives here)
  GARMIN_TIMESERIES_FILE    Full path for the output .xlsx file
  GARMIN_DATE_FROM          Start date (YYYY-MM-DD)
  GARMIN_DATE_TO            End date (YYYY-MM-DD)
"""

import json
import os
from datetime import date, datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════════
#  CONFIG — edit fallback values here, or set environment variables.
#  Environment variables always take priority over the values below.
# ══════════════════════════════════════════════════════════════════════════════

_BASE = Path(os.environ.get("GARMIN_OUTPUT_DIR", "~/garmin_data")).expanduser()

RAW_DIR     = _BASE / "raw"
OUTPUT_FILE = Path(os.environ.get("GARMIN_TIMESERIES_FILE",
                   str(_BASE / "garmin_timeseries.xlsx")))

# Date range — both required, format: "YYYY-MM-DD"
DATE_FROM = os.environ.get("GARMIN_DATE_FROM", "2026-03-01")
DATE_TO   = os.environ.get("GARMIN_DATE_TO",   "2026-03-16")

# ══════════════════════════════════════════════════════════════════════════════

# Metrics to export (True = include, False = skip)
METRICS = {
    "heart_rate":   True,
    "stress":       True,
    "spo2":         True,
    "body_battery": True,
    "respiration":  True,
}

# ══════════════════════════════════════════════════════════════════════════════
#  Styles
# ══════════════════════════════════════════════════════════════════════════════

HEADER_FILL = PatternFill("solid", start_color="1F3864")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT   = Font(name="Arial", size=10)
DATE_FONT   = Font(name="Arial", bold=True, size=10)
BORDER_SIDE = Side(style="thin", color="D0D0D0")
CELL_BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                     top=BORDER_SIDE,  bottom=BORDER_SIDE)

METRIC_COLORS = {
    "heart_rate":   "FFE8CC",
    "stress":       "E8FFE8",
    "spo2":         "DDEEFF",
    "body_battery": "FFF8CC",
    "respiration":  "F0E8FF",
}

METRIC_LABELS = {
    "heart_rate":   ("Heart Rate",   "bpm"),
    "stress":       ("Stress",       "level"),
    "spo2":         ("SpO2",         "%"),
    "body_battery": ("Body Battery", "level"),
    "respiration":  ("Respiration",  "brpm"),
}

CHART_COLORS = {
    "heart_rate":   "E85D24",
    "stress":       "1D9E75",
    "spo2":         "185FA5",
    "body_battery": "BA7517",
    "respiration":  "7F77DD",
}

# ══════════════════════════════════════════════════════════════════════════════
#  Data extraction from raw JSON
# ══════════════════════════════════════════════════════════════════════════════

def ts_to_time(ts) -> str:
    """Convert a Garmin timestamp (ms epoch or ISO string) to HH:MM."""
    if ts is None:
        return ""
    try:
        if isinstance(ts, (int, float)):
            return datetime.utcfromtimestamp(ts / 1000).strftime("%H:%M")
        return str(ts)[:16]
    except Exception:
        return str(ts)


def extract_heart_rate(raw: dict) -> list:
    """Returns list of (date, time, bpm)."""
    rows = []
    hr = raw.get("heart_rates") or {}
    values = hr.get("heartRateValues") or []
    day = raw.get("date", "")
    for item in values:
        if isinstance(item, (list, tuple)) and len(item) >= 2:
            ts, val = item[0], item[1]
        elif isinstance(item, dict):
            ts, val = item.get("startGMT") or item.get("timestamp"), item.get("heartRate") or item.get("value")
        else:
            continue
        if val is not None:
            rows.append((day, ts_to_time(ts), val))
    return rows


def extract_stress(raw: dict) -> list:
    """stress.stressValuesArray = [[ts, val], ...] with optional offset."""
    day    = raw.get("date", "")
    src    = raw.get("stress") or {}
    rows   = []
    if not isinstance(src, dict):
        return rows
    arr    = src.get("stressValuesArray") or []
    offset = src.get("stressChartValueOffset") or 0
    for item in arr:
        if isinstance(item, (list, tuple)) and len(item) >= 2:
            ts, val = item[0], item[1]
        elif isinstance(item, dict):
            ts  = item.get("startGMT") or item.get("timestamp")
            val = item.get("stressLevel") or item.get("value")
        else:
            continue
        try:
            v = float(val) - offset
            if v >= 0:
                rows.append((day, ts_to_time(ts), v))
        except (TypeError, ValueError):
            continue
    return rows


def extract_spo2(raw: dict) -> list:
    rows = []
    day = raw.get("date", "")
    spo2 = raw.get("spo2") or {}
    if isinstance(spo2, dict):
        values = spo2.get("spO2HourlyAverages") or spo2.get("continuousReadingDTOList") or []
    elif isinstance(spo2, list):
        values = spo2
    else:
        values = []
    for item in values:
        if isinstance(item, dict):
            ts  = item.get("startGMT") or item.get("timestamp")
            val = item.get("spO2Reading") or item.get("averageSpO2") or item.get("value")
        elif isinstance(item, (list, tuple)) and len(item) >= 2:
            ts, val = item[0], item[1]
        else:
            continue
        if val is not None:
            rows.append((day, ts_to_time(ts), val))
    return rows


def extract_body_battery(raw: dict) -> list:
    """stress.bodyBatteryValuesArray = [[ts, status, level, version], ...]"""
    day  = raw.get("date", "")
    rows = []
    src  = raw.get("stress") or {}
    arr  = src.get("bodyBatteryValuesArray") if isinstance(src, dict) else None
    if not arr:
        bb  = raw.get("body_battery") or {}
        arr = bb.get("bodyBatteryValuesArray") if isinstance(bb, dict) else None
    if not isinstance(arr, list):
        return rows
    for item in arr:
        if isinstance(item, (list, tuple)) and len(item) >= 3:
            ts, val = item[0], item[2]
        elif isinstance(item, dict):
            ts  = item.get("startGMT") or item.get("timestamp")
            val = item.get("bodyBatteryLevel") or item.get("value")
        else:
            continue
        try:
            rows.append((day, ts_to_time(ts), float(val)))
        except (TypeError, ValueError):
            continue
    return rows


def extract_respiration(raw: dict) -> list:
    rows = []
    day = raw.get("date", "")
    resp = raw.get("respiration") or {}
    if isinstance(resp, dict):
        values = resp.get("respirationValuesArray") or resp.get("respirationValues") or []
    elif isinstance(resp, list):
        values = resp
    else:
        values = []
    for item in values:
        if isinstance(item, dict):
            ts  = item.get("startGMT") or item.get("timestamp")
            val = item.get("respirationValue") or item.get("value")
        elif isinstance(item, (list, tuple)) and len(item) >= 2:
            ts, val = item[0], item[1]
        else:
            continue
        if val is not None:
            rows.append((day, ts_to_time(ts), val))
    return rows


EXTRACTORS = {
    "heart_rate":   extract_heart_rate,
    "stress":       extract_stress,
    "spo2":         extract_spo2,
    "body_battery": extract_body_battery,
    "respiration":  extract_respiration,
}

# ══════════════════════════════════════════════════════════════════════════════
#  Load raw files
# ══════════════════════════════════════════════════════════════════════════════

def load_raw_files() -> list[dict]:
    d_from = date.fromisoformat(DATE_FROM)
    d_to   = date.fromisoformat(DATE_TO)
    raws   = []
    for f in sorted(RAW_DIR.glob("garmin_raw_*.json")):
        try:
            d = date.fromisoformat(f.stem.replace("garmin_raw_", ""))
        except ValueError:
            continue
        if d_from <= d <= d_to:
            with open(f, encoding="utf-8") as fp:
                raws.append(json.load(fp))
    print(f"  {len(raws)} raw files loaded ({DATE_FROM} → {DATE_TO})")
    return raws

# ══════════════════════════════════════════════════════════════════════════════
#  Excel builder
# ══════════════════════════════════════════════════════════════════════════════

def write_metric_sheets(wb: Workbook, metric: str, all_rows: list):
    label, unit = METRIC_LABELS[metric]
    fill_color  = METRIC_COLORS[metric]
    chart_color = CHART_COLORS[metric]

    data_title  = f"{label} - Data"
    chart_title = f"{label} - Chart"

    # ── Data sheet ────────────────────────────────────────────────────────────
    ws = wb.create_sheet(data_title)
    ws.freeze_panes = "A2"

    headers = ["Date", "Time", unit]
    widths  = [12, 8, 10]

    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(1, col, h)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border    = CELL_BORDER
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 24
    row_fill = PatternFill("solid", start_color=fill_color)

    for row_idx, (day, time, val) in enumerate(all_rows, start=2):
        ws.cell(row_idx, 1, day).font   = DATE_FONT
        ws.cell(row_idx, 1).alignment   = Alignment(horizontal="center")
        ws.cell(row_idx, 1).border      = CELL_BORDER
        ws.cell(row_idx, 2, time).font  = DATA_FONT
        ws.cell(row_idx, 2).alignment   = Alignment(horizontal="center")
        ws.cell(row_idx, 2).border      = CELL_BORDER
        ws.cell(row_idx, 3, val).font   = DATA_FONT
        ws.cell(row_idx, 3).fill        = row_fill
        ws.cell(row_idx, 3).alignment   = Alignment(horizontal="center")
        ws.cell(row_idx, 3).border      = CELL_BORDER
        if isinstance(val, float):
            ws.cell(row_idx, 3).number_format = "0.0"

    # ── Chart sheet ───────────────────────────────────────────────────────────
    wc = wb.create_sheet(chart_title)

    chart = LineChart()
    chart.title       = f"{label} ({DATE_FROM} – {DATE_TO})"
    chart.style       = 10
    chart.y_axis.title = unit
    chart.x_axis.title = "Measurement"
    chart.width        = 30
    chart.height       = 18

    n = len(all_rows)
    if n > 0:
        data_ref = Reference(ws, min_col=3, min_row=1, max_row=n + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.series[0].graphicalProperties.line.solidFill = chart_color
        chart.series[0].graphicalProperties.line.width     = 15000  # 1.5pt in EMU

    wc.add_chart(chart, "A1")
    print(f"  ✓ {label}: {len(all_rows)} data points")


# ══════════════════════════════════════════════════════════════════════════════
#  Main
# ══════════════════════════════════════════════════════════════════════════════

def main():
    print("Garmin → Timeseries Excel export")
    print(f"  Source:  {RAW_DIR}")
    print(f"  Output:  {OUTPUT_FILE}")
    print(f"  Range:   {DATE_FROM} → {DATE_TO}")

    if not RAW_DIR.exists():
        print(f"  ERROR: folder not found: {RAW_DIR}")
        return

    raws = load_raw_files()
    if not raws:
        print("  No raw files found for the given date range.")
        return

    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    for metric, enabled in METRICS.items():
        if not enabled:
            continue
        extractor = EXTRACTORS[metric]
        all_rows  = []
        for raw in raws:
            all_rows.extend(extractor(raw))
        write_metric_sheets(wb, metric, all_rows)

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_FILE)
    print(f"  ✓ Saved: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
